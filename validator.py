# validator.py
from __future__ import annotations

from curses import meta
import json
import re
import uuid
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional, Union
from zipfile import ZipFile

# =========================================================
# Error catalog (Day11)
# =========================================================

ERROR_CATALOG: Dict[str, Dict[str, str]] = {
    "XML_PARSE_ERROR": {
        "title": "XML parsing failed",
        "explanation": "The uploaded file is not a valid XML or is truncated.",
        "action": "Re-download the XML or ensure the file is a complete ISO20022 XML message.",
    },
    "MESSAGE_TYPE_NOT_DETECTED": {
        "title": "Message type not detected",
        "explanation": "The XML namespace does not contain a recognizable ISO20022 message identifier.",
        "action": "Ensure the XML uses a standard ISO20022 namespace or select a ruleset manually.",
    },
    "RULESET_NOT_FOUND": {
        "title": "Ruleset not found",
        "explanation": "No matching ruleset file exists under the rules/ directory.",
        "action": "Create an appropriate ruleset (e.g., rules/pacs.008-core.json) or select another ruleset.",
    },
    "RULESET_MAPPING_MISSING": {
        "title": "Ruleset mapping missing",
        "explanation": "Batch validation requires a ruleset selection per detected message type.",
        "action": "Select a ruleset for each message type in the mapping section.",
    },
    # Common required fields (examples)
    "MISSING_MSGID": {
        "title": "Missing Message Identification (MsgId)",
        "explanation": "MsgId uniquely identifies the message in the group header for tracking and reconciliation.",
        "action": "Populate GrpHdr/MsgId with a unique identifier.",
    },
    "MISSING_TXID": {
        "title": "Missing Transaction Identifier",
        "explanation": "A transaction identifier is required to uniquely reference the payment instruction.",
        "action": "Populate one of: PmtId/TxId, PmtId/EndToEndId, or PmtId/InstrId (per your scheme).",
    },
    "MISSING_AMOUNT": {
        "title": "Missing Instructed Amount",
        "explanation": "The instructed amount is required to execute the transfer.",
        "action": "Populate CdtTrfTxInf/Amt/InstdAmt with a numeric value and currency attribute.",
    },
    "MISSING_CURRENCY": {
        "title": "Missing Currency (Ccy attribute)",
        "explanation": "Currency must be provided as the Ccy attribute on InstdAmt.",
        "action": "Set Ccy attribute on CdtTrfTxInf/Amt/InstdAmt (e.g., Ccy='EUR').",
    },
    "MISSING_UETR": {
        "title": "UETR missing",
        "explanation": "UETR helps uniquely track cross-border payments and is commonly used in gpi flows.",
        "action": "Populate CdtTrfTxInf/PmtId/UETR with a UUID if your scheme supports it.",
    },
}

# =========================================================
# Rulesets + License filtering (Day19/23/24)
# =========================================================


def list_rulesets(rules_dir: str = "rules") -> List[str]:
    """Return ruleset stems from rules/*.json, e.g. ['pacs.008-core', 'pacs.008-gpi']"""
    p = Path(rules_dir)
    if not p.exists():
        return []
    return sorted([f.stem for f in p.glob("*.json")])


def ruleset_to_message_type(ruleset: str) -> str:
    """e.g. pacs.008-core -> pacs.008"""
    if "-" not in ruleset:
        return ruleset
    return ruleset.split("-", 1)[0]


def find_rulesets_for_message_type(message_type: str, rules_dir: str = "rules") -> List[str]:
    all_sets = list_rulesets(rules_dir)
    return [rs for rs in all_sets if ruleset_to_message_type(rs) == message_type]


def choose_default_ruleset(rulesets: List[str]) -> Optional[str]:
    """Prefer *-core, otherwise first."""
    for rs in rulesets:
        if rs.endswith("-core"):
            return rs
    return rulesets[0] if rulesets else None


def is_enterprise_ruleset(ruleset: str) -> bool:
    """Convention: only *-core is community."""
    return not ruleset.endswith("-core")


def filter_rulesets_by_license(all_rulesets: List[str], is_enterprise_enabled: bool) -> List[str]:
    if is_enterprise_enabled:
        return all_rulesets
    return [rs for rs in all_rulesets if rs.endswith("-core")]


def validate_rules_schema(rules: dict) -> dict:
    """
    Minimal schema validation (Day18).
    Requires meta.id/meta.version/meta.released, message_type, required_fields dict.
    """
    meta = rules.get("meta") or {}
    required_meta_keys = ["id", "version", "released"]
    missing = [k for k in required_meta_keys if not meta.get(k)]
    if missing:
        raise ValueError(f"Rules meta missing keys: {missing}. Please add meta.{missing} in rules json.")

    if not rules.get("message_type"):
        raise ValueError("Rules missing 'message_type'.")

    if not isinstance(rules.get("required_fields", {}), dict):
        raise ValueError("Rules 'required_fields' must be an object/dict.")

    # optional_fields/format_rules allowed to be empty
    return meta


def load_rules_by_ruleset(ruleset: str, rules_dir: str = "rules") -> dict:
    """Load rules/{ruleset}.json"""
    rules_file = Path(rules_dir) / f"{ruleset}.json"
    if not rules_file.exists():
        raise FileNotFoundError(f"Ruleset not found: {rules_file}")

    with open(rules_file, "r", encoding="utf-8") as f:
        rules = json.load(f)

    validate_rules_schema(rules)
    return rules

    meta.setdefault("profile", "unknown")
    meta.setdefault("tier", "unknown")
    return meta

# =========================================================
# XML helpers
# =========================================================


def strip_namespace(root: ET.Element) -> ET.Element:
    for elem in root.iter():
        if "}" in elem.tag:
            elem.tag = elem.tag.split("}", 1)[1]
    return root


def safe_find_text(root: ET.Element, path: str) -> Optional[str]:
    node = root.find(path)
    if node is not None and node.text is not None:
        text = node.text.strip()
        return text if text != "" else None
    return None


# =========================================================
# Auto-detect message type from XML (Day15/20)
# =========================================================


def detect_message_type_from_xml(xml_bytes: bytes) -> Optional[str]:
    """
    Detect message type from namespace:
    urn:iso:std:iso:20022:tech:xsd:pacs.008.001.08 -> pacs.008
    """
    try:
        root = ET.fromstring(xml_bytes)
    except ET.ParseError:
        return None

    tag = root.tag  # maybe {namespace}Document
    if not tag or "}" not in tag:
        return None

    ns = tag.split("}", 1)[0].strip("{")
    m = re.search(r"\b([a-z]{4})\.(\d{3})\.\d{3}\.\d{2}\b", ns)
    if not m:
        return None

    family = m.group(1)  # pacs/pain/camt
    num = m.group(2)     # 008/009/053...
    return f"{family}.{num}"


# =========================================================
# Field extraction with multi-xpath fallback (Day10)
# =========================================================

PathSpec = Union[str, List[str]]


def first_match_text(root: ET.Element, paths: PathSpec) -> Optional[str]:
    if isinstance(paths, str):
        paths = [paths]

    for p in paths:
        node = root.find(p)
        if node is None:
            continue

        if node.text is not None:
            text = node.text.strip()
            if text != "":
                return text

    return None


def extract_fields(root: ET.Element, field_paths: Dict[str, PathSpec]) -> Dict[str, Optional[str]]:
    result: Dict[str, Optional[str]] = {}

    for field_name, paths in field_paths.items():
        # Currency special: from attribute Ccy on amount node(s)
        if field_name.lower() == "currency":
            if isinstance(paths, str):
                paths = [paths]
            ccy = None
            for p in paths:
                node = root.find(p)
                if node is not None and node.get("Ccy"):
                    ccy = node.get("Ccy")
                    break
            result[field_name] = ccy
            continue

        result[field_name] = first_match_text(root, paths)

    return result


def validate_required_fields(fields: Dict[str, Any], required_field_names) -> List[str]:
    missing = []
    for name in required_field_names:
        value = fields.get(name)
        if value is None or (isinstance(value, str) and value.strip() == ""):
            missing.append(name)
    return missing


# =========================================================
# Format validation (Day16/17)
# =========================================================


def validate_format_rules(parsed_fields: Dict[str, Any], format_rules: Dict[str, dict]) -> List[dict]:
    """
    Validate fields by format_rules.
    Only validates fields that exist (not None). Missing is handled by required_fields.
    Supported types: decimal, currency_code, uuid
    """
    issues: List[dict] = []

    for field, rule in (format_rules or {}).items():
        value = parsed_fields.get(field)

        if value is None:
            continue

        rtype = rule.get("type")

        # ---- decimal ----
        if rtype == "decimal":
            try:
                dec = Decimal(str(value))
            except (InvalidOperation, ValueError):
                issues.append({
                    "code": f"INVALID_{field.upper()}_FORMAT",
                    "field": field,
                    "level": "ERROR",
                    "message": f"{field} must be a valid decimal number.",
                    "title": f"Invalid {field} format",
                    "explanation": f"The value '{value}' is not a valid decimal.",
                    "action": f"Provide a numeric value for {field}, e.g., 1000.00",
                })
                continue

            min_v = rule.get("min")
            if min_v is not None:
                try:
                    if dec <= Decimal(str(min_v)):
                        issues.append({
                            "code": f"INVALID_{field.upper()}_RANGE",
                            "field": field,
                            "level": "ERROR",
                            "message": f"{field} must be greater than {min_v}.",
                            "title": f"{field} out of range",
                            "explanation": f"The value {value} is not > {min_v}.",
                            "action": f"Provide a positive amount for {field}.",
                        })
                except (InvalidOperation, ValueError):
                    pass

        # ---- currency_code ----
        elif rtype == "currency_code":
            v = str(value).strip()
            if not re.fullmatch(r"[A-Z]{3}", v):
                issues.append({
                    "code": "INVALID_CURRENCY_CODE",
                    "field": field,
                    "level": "ERROR",
                    "message": "Currency must be a 3-letter uppercase ISO code (e.g., EUR).",
                    "title": "Invalid currency code",
                    "explanation": f"The value '{value}' is not a valid 3-letter currency code.",
                    "action": "Set the Ccy attribute to a valid ISO 4217 code (EUR, USD, GBP, etc.).",
                })

        # ---- uuid (UETR) ----
        elif rtype == "uuid":
            v = str(value).strip()
            try:
                uuid.UUID(v)
            except ValueError:
                issues.append({
                    "code": f"INVALID_{field.upper()}",
                    "field": field,
                    "level": "ERROR",
                    "message": f"{field} must be a valid UUID (e.g., 550e8400-e29b-41d4-a716-446655440000).",
                    "title": f"Invalid {field}",
                    "explanation": f"The value '{value}' is not a valid UUID format.",
                    "action": f"Provide a valid UUID for {field} (36 characters with hyphens).",
                })

    return issues


# =========================================================
# Validation result builder (Day6/11)
# =========================================================


def build_validation_result(fields: Dict[str, Any], missing_required: List[str], message_type: str) -> dict:
    result = {
        "message_type": message_type,
        "status": "VALID",
        "errors": [],
        "warnings": [],
        "parsed_fields": fields,
    }

    if missing_required:
        result["status"] = "INVALID"
        for field in missing_required:
            code = f"MISSING_{field.upper()}"
            info = ERROR_CATALOG.get(code, {})
            result["errors"].append({
                "code": code,
                "field": field,
                "level": "ERROR",
                "message": f"{field} is mandatory for {message_type}",
                "title": info.get("title"),
                "explanation": info.get("explanation"),
                "action": info.get("action"),
            })

    return result


# =========================================================
# Main single-file validation
# =========================================================


def validate_iso20022_xml(xml_bytes: bytes, rules: dict) -> dict:
    """
    Input: XML bytes
    Output: validation_result dict
    """
    # Parse XML
    try:
        root = ET.fromstring(xml_bytes)
    except ET.ParseError as e:
        info = ERROR_CATALOG.get("XML_PARSE_ERROR", {})
        return {
            "message_type": rules.get("message_type", "unknown"),
            "status": "INVALID",
            "errors": [{
                "code": "XML_PARSE_ERROR",
                "field": None,
                "level": "ERROR",
                "message": f"XML parse error: {e}",
                "title": info.get("title"),
                "explanation": info.get("explanation"),
                "action": info.get("action"),
            }],
            "warnings": [],
            "parsed_fields": {},
            "optional_fields": {},
            "rules_meta": rules.get("meta", {}),
        }

    root = strip_namespace(root)

    message_type = rules.get("message_type", "unknown")
    required_paths = rules.get("required_fields", {}) or {}
    optional_paths = rules.get("optional_fields", {}) or {}

    fields_required = extract_fields(root, required_paths)
    missing_required = validate_required_fields(fields_required, required_paths.keys())
    validation_result = build_validation_result(fields_required, missing_required, message_type)

    # Optional fields for display / warning checks
    if optional_paths:
        validation_result["optional_fields"] = extract_fields(root, optional_paths)
    else:
        validation_result["optional_fields"] = {}

    # Format rules (Day16/17)
    format_rules = rules.get("format_rules", {}) or {}
    format_issues = validate_format_rules(validation_result["parsed_fields"], format_rules)
    if format_issues:
        validation_result["errors"].extend(format_issues)
        validation_result["status"] = "INVALID"

    # UETR missing warning (Day17 add-on)
    # If format_rules includes UETR uuid check, but UETR missing -> warning
    if "UETR" in format_rules and format_rules.get("UETR", {}).get("type") == "uuid":
        uetr_val = None
        if isinstance(validation_result.get("optional_fields"), dict):
            uetr_val = validation_result["optional_fields"].get("UETR")
        if uetr_val is None:
            info = ERROR_CATALOG.get("MISSING_UETR", {})
            validation_result["warnings"].append({
                "code": "MISSING_UETR",
                "field": "UETR",
                "level": "WARNING",
                "message": "UETR is recommended for tracking (SWIFT gpi), but it is missing.",
                "title": info.get("title"),
                "explanation": info.get("explanation"),
                "action": info.get("action"),
            })

    # Attach rules meta (Day18)
    validation_result["rules_meta"] = rules.get("meta", {})

    return validation_result


# =========================================================
# Excel report (single-file) (Day12/18/25)
# =========================================================


def _autofit_columns(ws, max_col: int):
    from openpyxl.utils import get_column_letter

    for col in range(1, max_col + 1):
        col_letter = get_column_letter(col)
        max_len = 0
        for cell in ws[col_letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


def build_excel_report_bytes(validation_result: dict) -> bytes:
    """
    Create an Excel report for a single validation_result.
    Sheets: Summary / Parsed Fields / Issues
    """
    from openpyxl import Workbook

    wb = Workbook()

    # ===== Sheet 1: Summary =====
    ws1 = wb.active
    ws1.title = "Summary"

    message_type = validation_result.get("message_type", "unknown")
    status = validation_result.get("status", "unknown")
    errors = validation_result.get("errors", []) or []
    warnings = validation_result.get("warnings", []) or []
    rules_meta = validation_result.get("rules_meta", {}) or {}

    # Day25 fields (optional, populated by app.py)
    license_level = validation_result.get("license_level")
    ruleset_used = validation_result.get("ruleset_used")
    ruleset_tier = validation_result.get("ruleset_tier")

    summary_rows = [
        ("Message Type", message_type),
        ("Status", status),
        ("License Level", license_level),
        ("Ruleset Used", ruleset_used),
        ("Ruleset Tier", ruleset_tier),
        ("Ruleset ID", rules_meta.get("id")),
        ("Ruleset Version", rules_meta.get("version")),
        ("Ruleset Released", rules_meta.get("released")),
        ("Error Count", len(errors)),
        ("Warning Count", len(warnings)),
    ]

    ws1.append(["Key", "Value"])
    for k, v in summary_rows:
        ws1.append([k, v])

    _autofit_columns(ws1, 2)

    # ===== Sheet 2: Parsed Fields =====
    ws2 = wb.create_sheet("Parsed Fields")
    ws2.append(["Field", "Value"])

    parsed = validation_result.get("parsed_fields", {}) or {}
    for k, v in parsed.items():
        ws2.append([k, v])

    optional = validation_result.get("optional_fields")
    if isinstance(optional, dict) and optional:
        ws2.append([])
        ws2.append(["Optional Field", "Value"])
        for k, v in optional.items():
            ws2.append([k, v])

    _autofit_columns(ws2, 2)

    # ===== Sheet 3: Issues =====
    ws3 = wb.create_sheet("Issues")
    ws3.append(["Level", "Code", "Field", "Message", "Title", "Explanation", "Suggested Fix"])

    def add_issue(level: str, issue: dict):
        ws3.append([
            level,
            issue.get("code"),
            issue.get("field"),
            issue.get("message"),
            issue.get("title"),
            issue.get("explanation"),
            issue.get("action"),
        ])

    for e in errors:
        add_issue("ERROR", e)

    for w in warnings:
        add_issue("WARNING", w)

    _autofit_columns(ws3, 7)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


# =========================================================
# Batch ZIP validation (Day21/22)
# =========================================================


def validate_zip_bytes_auto(zip_bytes: bytes, rules_dir: str = "rules") -> List[dict]:
    """
    Auto-detect message_type per file, choose default ruleset (*-core preferred), validate.
    Returns list of:
      {file_name, detected_type, ruleset_used, rules_meta, validation_result}
    """
    results: List[dict] = []
    bio = BytesIO(zip_bytes)

    with ZipFile(bio, "r") as z:
        for name in z.namelist():
            if name.endswith("/") or not name.lower().endswith(".xml"):
                continue

            xml_b = z.read(name)
            detected = detect_message_type_from_xml(xml_b)

            if not detected:
                info = ERROR_CATALOG.get("MESSAGE_TYPE_NOT_DETECTED", {})
                vr = {
                    "message_type": None,
                    "status": "INVALID",
                    "errors": [{
                        "code": "MESSAGE_TYPE_NOT_DETECTED",
                        "field": None,
                        "level": "ERROR",
                        "message": "Unable to detect ISO20022 message type from XML namespace.",
                        "title": info.get("title"),
                        "explanation": info.get("explanation"),
                        "action": info.get("action"),
                    }],
                    "warnings": [],
                    "parsed_fields": {},
                    "optional_fields": {},
                    "rules_meta": {},
                }
                results.append({
                    "file_name": name,
                    "detected_type": None,
                    "ruleset_used": None,
                    "rules_meta": {},
                    "validation_result": vr,
                })
                continue

            matched = find_rulesets_for_message_type(detected, rules_dir)
            if not matched:
                info = ERROR_CATALOG.get("RULESET_NOT_FOUND", {})
                vr = {
                    "message_type": detected,
                    "status": "INVALID",
                    "errors": [{
                        "code": "RULESET_NOT_FOUND",
                        "field": None,
                        "level": "ERROR",
                        "message": f"No rulesets found for detected message type {detected}.",
                        "title": info.get("title"),
                        "explanation": info.get("explanation"),
                        "action": info.get("action"),
                    }],
                    "warnings": [],
                    "parsed_fields": {},
                    "optional_fields": {},
                    "rules_meta": {},
                }
                results.append({
                    "file_name": name,
                    "detected_type": detected,
                    "ruleset_used": None,
                    "rules_meta": {},
                    "validation_result": vr,
                })
                continue

            default_rs = choose_default_ruleset(matched)
            rules = load_rules_by_ruleset(default_rs, rules_dir)

            vr = validate_iso20022_xml(xml_b, rules)
            vr["message_type"] = detected
            vr["rules_meta"] = rules.get("meta", {})

            results.append({
                "file_name": name,
                "detected_type": detected,
                "ruleset_used": default_rs,
                "rules_meta": rules.get("meta", {}),
                "validation_result": vr,
            })

    return results


def validate_zip_bytes_with_ruleset_mapping(
    zip_bytes: bytes,
    ruleset_mapping: Dict[str, str],
    rules_dir: str = "rules",
) -> List[dict]:
    """
    Validate ZIP using mapping: { 'pacs.008': 'pacs.008-gpi', ... }
    Returns list of:
      {file_name, detected_type, ruleset_used, rules_meta, validation_result}
    """
    results: List[dict] = []
    bio = BytesIO(zip_bytes)

    with ZipFile(bio, "r") as z:
        for name in z.namelist():
            if name.endswith("/") or not name.lower().endswith(".xml"):
                continue

            xml_b = z.read(name)
            detected = detect_message_type_from_xml(xml_b)

            # ① not detected
            if not detected:
                info = ERROR_CATALOG.get("MESSAGE_TYPE_NOT_DETECTED", {})
                vr = {
                    "message_type": None,
                    "status": "INVALID",
                    "errors": [{
                        "code": "MESSAGE_TYPE_NOT_DETECTED",
                        "field": None,
                        "level": "ERROR",
                        "message": "Unable to detect ISO20022 message type from XML namespace.",
                        "title": info.get("title"),
                        "explanation": info.get("explanation"),
                        "action": info.get("action"),
                    }],
                    "warnings": [],
                    "parsed_fields": {},
                    "optional_fields": {},
                    "rules_meta": {},
                }
                results.append({
                    "file_name": name,
                    "detected_type": None,
                    "ruleset_used": None,
                    "rules_meta": {},
                    "validation_result": vr,
                })
                continue

            chosen_ruleset = ruleset_mapping.get(detected)
            if not chosen_ruleset:
                info = ERROR_CATALOG.get("RULESET_MAPPING_MISSING", {})
                vr = {
                    "message_type": detected,
                    "status": "INVALID",
                    "errors": [{
                        "code": "RULESET_MAPPING_MISSING",
                        "field": None,
                        "level": "ERROR",
                        "message": f"No ruleset selected for detected message type {detected}.",
                        "title": info.get("title"),
                        "explanation": info.get("explanation"),
                        "action": info.get("action"),
                    }],
                    "warnings": [],
                    "parsed_fields": {},
                    "optional_fields": {},
                    "rules_meta": {},
                }
                results.append({
                    "file_name": name,
                    "detected_type": detected,
                    "ruleset_used": None,
                    "rules_meta": {},
                    "validation_result": vr,
                })
                continue

            try:
                rules = load_rules_by_ruleset(chosen_ruleset, rules_dir)
            except FileNotFoundError:
                info = ERROR_CATALOG.get("RULESET_NOT_FOUND", {})
                vr = {
                    "message_type": detected,
                    "status": "INVALID",
                    "errors": [{
                        "code": "RULESET_NOT_FOUND",
                        "field": None,
                        "level": "ERROR",
                        "message": f"Ruleset file not found: {chosen_ruleset}.json",
                        "title": info.get("title"),
                        "explanation": info.get("explanation"),
                        "action": info.get("action"),
                    }],
                    "warnings": [],
                    "parsed_fields": {},
                    "optional_fields": {},
                    "rules_meta": {},
                }
                results.append({
                    "file_name": name,
                    "detected_type": detected,
                    "ruleset_used": chosen_ruleset,
                    "rules_meta": {},
                    "validation_result": vr,
                })
                continue

            vr = validate_iso20022_xml(xml_b, rules)
            vr["message_type"] = detected
            vr["rules_meta"] = rules.get("meta", {})

            results.append({
                "file_name": name,
                "detected_type": detected,
                "ruleset_used": chosen_ruleset,
                "rules_meta": rules.get("meta", {}),
                "validation_result": vr,
            })

    return results


# =========================================================
# Batch Excel report (Day21/22/25)
# =========================================================


def build_batch_excel_report_bytes(batch_results: List[dict]) -> bytes:
    """
    Excel report for batch results.
    Sheets: Summary / Issues / Parsed Fields
    """
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    def autofit(ws, max_col):
        for col in range(1, max_col + 1):
            letter = get_column_letter(col)
            max_len = 0
            for cell in ws[letter]:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[letter].width = min(max_len + 2, 60)

    wb = Workbook()

    # ===== Sheet 1: Summary =====
    ws1 = wb.active
    ws1.title = "Summary"

    # Day25: include License Level & Ruleset Tier if present in each item
    ws1.append([
        "File",
        "Detected Type",
        "License Level",
        "Ruleset Used",
        "Ruleset Tier",
        "Ruleset Version",
        "Status",
        "Error Count",
        "Warning Count",
    ])

    for item in batch_results:
        fn = item.get("file_name")
        detected = item.get("detected_type")
        ruleset_used = item.get("ruleset_used")
        meta = item.get("rules_meta", {}) or {}
        vr = item.get("validation_result", {}) or {}

        license_level = item.get("license_level")  # set by app.py (Day25)
        ruleset_tier = (ruleset_used.split("-", 1)[1] if ruleset_used and "-" in ruleset_used else "unknown")

        ws1.append([
            fn,
            detected,
            license_level,
            ruleset_used,
            ruleset_tier,
            meta.get("version"),
            vr.get("status", "unknown"),
            len(vr.get("errors", []) or []),
            len(vr.get("warnings", []) or []),
        ])

    autofit(ws1, 9)

    # ===== Sheet 2: Issues =====
    ws2 = wb.create_sheet("Issues")
    ws2.append([
        "File",
        "Detected Type",
        "Ruleset Used",
        "Level",
        "Code",
        "Field",
        "Message",
        "Title",
        "Explanation",
        "Suggested Fix",
    ])

    for item in batch_results:
        fn = item.get("file_name")
        detected = item.get("detected_type")
        ruleset_used = item.get("ruleset_used")
        vr = item.get("validation_result", {}) or {}

        for e in vr.get("errors", []) or []:
            ws2.append([
                fn, detected, ruleset_used, "ERROR",
                e.get("code"), e.get("field"), e.get("message"),
                e.get("title"), e.get("explanation"), e.get("action")
            ])

        for w in vr.get("warnings", []) or []:
            ws2.append([
                fn, detected, ruleset_used, "WARNING",
                w.get("code"), w.get("field"), w.get("message"),
                w.get("title"), w.get("explanation"), w.get("action")
            ])

    autofit(ws2, 10)

    # ===== Sheet 3: Parsed Fields =====
    ws3 = wb.create_sheet("Parsed Fields")
    ws3.append(["File", "Detected Type", "Ruleset Used", "Field", "Value"])

    for item in batch_results:
        fn = item.get("file_name")
        detected = item.get("detected_type")
        ruleset_used = item.get("ruleset_used")
        vr = item.get("validation_result", {}) or {}
        parsed = vr.get("parsed_fields", {}) or {}

        for k, v in parsed.items():
            ws3.append([fn, detected, ruleset_used, k, v])

    autofit(ws3, 5)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
